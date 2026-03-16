from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import csv
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = ROOT / "templates"
DATA_DIR = ROOT / "data"
CUSTOM_RESULTS_PATH = DATA_DIR / "custom_results.csv"
NORMALIZED_EXPORT_PATH = DATA_DIR / "normalized_results.csv"

DATE_FORMAT = "%A, %d %b %Y"
DRAW_TYPES = {"midday", "evening"}


@dataclass(frozen=True)
class DrawRecord:
    draw_type: str
    draw_date: date
    draw_number: int | None
    digits: tuple[int, int, int]
    source: str

    @property
    def number(self) -> str:
        return "-".join(str(digit) for digit in self.digits)

    def to_dict(self) -> dict[str, object]:
        payload = asdict(self)
        payload["draw_date"] = self.draw_date.isoformat()
        payload["number"] = self.number
        return payload


def parse_number(raw_value: str) -> tuple[int, int, int]:
    digits = [int(char) for char in re.findall(r"\d", raw_value or "")]
    if len(digits) != 3:
        raise ValueError("Winning number must contain exactly 3 digits.")
    return digits[0], digits[1], digits[2]


def parse_draw_date(raw_value: str) -> date:
    return datetime.strptime(raw_value.strip(), DATE_FORMAT).date()


def _load_workbook_records(path: Path) -> list[DrawRecord]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = [row[0] for row in worksheet.iter_rows(values_only=True)]

    records: list[DrawRecord] = []
    index = 0
    while index < len(rows):
        label = rows[index]
        if isinstance(label, str) and label.strip().lower() in DRAW_TYPES:
            draw_type = label.strip().lower()
            try:
                draw_date = parse_draw_date(str(rows[index + 2]))
                draw_number = _extract_draw_number(rows[index + 3])
                digits = tuple(int(rows[index + offset]) for offset in (5, 6, 7))
            except (IndexError, TypeError, ValueError):
                index += 1
                continue

            records.append(
                DrawRecord(
                    draw_type=draw_type,
                    draw_date=draw_date,
                    draw_number=draw_number,
                    digits=digits,
                    source=path.name,
                )
            )
            index += 8
            continue

        index += 1

    return records


def _extract_draw_number(raw_value: object) -> int | None:
    if raw_value is None:
        return None
    match = re.search(r"(\d+)", str(raw_value))
    return int(match.group(1)) if match else None


def load_all_records() -> pd.DataFrame:
    DATA_DIR.mkdir(exist_ok=True)

    records: list[DrawRecord] = []
    for path in sorted(TEMPLATES_DIR.glob("*.xlsx")):
        records.extend(_load_workbook_records(path))

    if CUSTOM_RESULTS_PATH.exists():
        records.extend(_load_custom_records(CUSTOM_RESULTS_PATH))

    deduped = {}
    for record in records:
        key = (
            record.draw_type,
            record.draw_date.isoformat(),
            record.draw_number,
            record.digits,
        )
        deduped[key] = record

    frame = pd.DataFrame(record.to_dict() for record in deduped.values())
    if frame.empty:
        return frame

    frame["draw_date"] = pd.to_datetime(frame["draw_date"])
    frame["sort_draw_number"] = frame["draw_number"].fillna(-1)
    frame = frame.sort_values(
        by=["draw_type", "draw_date", "sort_draw_number"],
        ascending=[True, False, False],
    ).reset_index(drop=True)
    frame = frame.drop(columns=["sort_draw_number"])
    frame.to_csv(NORMALIZED_EXPORT_PATH, index=False)
    return frame


def _load_custom_records(path: Path) -> list[DrawRecord]:
    records: list[DrawRecord] = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            records.append(
                DrawRecord(
                    draw_type=str(row["draw_type"]).strip().lower(),
                    draw_date=datetime.strptime(row["draw_date"], "%Y-%m-%d").date(),
                    draw_number=int(row["draw_number"]) if row["draw_number"] else None,
                    digits=parse_number(str(row["number"])),
                    source=str(row.get("source") or "manual"),
                )
            )
    return records


def append_custom_record(
    draw_type: str,
    draw_date: date,
    winning_number: str,
    draw_number: int | None = None,
) -> DrawRecord:
    normalized_type = draw_type.strip().lower()
    if normalized_type not in DRAW_TYPES:
        raise ValueError("Draw type must be either midday or evening.")

    record = DrawRecord(
        draw_type=normalized_type,
        draw_date=draw_date,
        draw_number=draw_number,
        digits=parse_number(winning_number),
        source="manual",
    )

    DATA_DIR.mkdir(exist_ok=True)
    existing_records = [
        DrawRecord(
            draw_type=str(row["draw_type"]).strip().lower(),
            draw_date=pd.Timestamp(row["draw_date"]).date(),
            draw_number=int(row["draw_number"]) if pd.notna(row["draw_number"]) else None,
            digits=parse_number(str(row["number"])),
            source=str(row["source"]),
        )
        for _, row in load_all_records().iterrows()
    ]
    already_exists = any(_record_equals(record, existing) for existing in existing_records)
    if already_exists:
        raise ValueError("This exact result already exists in the dataset.")

    write_header = not CUSTOM_RESULTS_PATH.exists()
    with CUSTOM_RESULTS_PATH.open("a", newline="", encoding="utf-8") as handle:
        fieldnames = ["draw_type", "draw_date", "draw_number", "number", "source"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow(
            {
                "draw_type": record.draw_type,
                "draw_date": record.draw_date.isoformat(),
                "draw_number": record.draw_number or "",
                "number": record.number,
                "source": record.source,
            }
        )

    return record


def list_custom_records() -> pd.DataFrame:
    if not CUSTOM_RESULTS_PATH.exists():
        return pd.DataFrame(columns=["entry_id", "draw_type", "draw_date", "draw_number", "number", "source"])

    frame = pd.read_csv(CUSTOM_RESULTS_PATH)
    if frame.empty:
        return pd.DataFrame(columns=["entry_id", "draw_type", "draw_date", "draw_number", "number", "source"])

    frame = frame.fillna({"draw_number": "", "source": "manual"})
    frame.insert(0, "entry_id", frame.index.astype(int))
    return frame


def update_custom_record(
    entry_id: int,
    draw_type: str,
    draw_date: date,
    winning_number: str,
    draw_number: int | None = None,
) -> DrawRecord:
    records = list(_load_custom_records_if_present())
    if entry_id < 0 or entry_id >= len(records):
        raise ValueError("Manual entry not found.")

    normalized_type = draw_type.strip().lower()
    if normalized_type not in DRAW_TYPES:
        raise ValueError("Draw type must be either midday or evening.")

    updated_record = DrawRecord(
        draw_type=normalized_type,
        draw_date=draw_date,
        draw_number=draw_number,
        digits=parse_number(winning_number),
        source="manual",
    )

    existing_records = [
        DrawRecord(
            draw_type=str(row["draw_type"]).strip().lower(),
            draw_date=pd.Timestamp(row["draw_date"]).date(),
            draw_number=int(row["draw_number"]) if pd.notna(row["draw_number"]) else None,
            digits=parse_number(str(row["number"])),
            source=str(row["source"]),
        )
        for _, row in load_all_records().iterrows()
    ]
    duplicate_exists = any(
        _record_equals(updated_record, existing)
        for existing in existing_records
        if existing != records[entry_id]
    )
    if duplicate_exists:
        raise ValueError("This exact result already exists in the dataset.")

    records[entry_id] = updated_record
    _write_custom_records(records)
    return updated_record


def delete_custom_record(entry_id: int) -> None:
    records = list(_load_custom_records_if_present())
    if entry_id < 0 or entry_id >= len(records):
        raise ValueError("Manual entry not found.")

    del records[entry_id]
    _write_custom_records(records)


def _load_custom_records_if_present() -> Iterable[DrawRecord]:
    if not CUSTOM_RESULTS_PATH.exists():
        return []
    return _load_custom_records(CUSTOM_RESULTS_PATH)


def _write_custom_records(records: Iterable[DrawRecord]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    fieldnames = ["draw_type", "draw_date", "draw_number", "number", "source"]

    with CUSTOM_RESULTS_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "draw_type": record.draw_type,
                    "draw_date": record.draw_date.isoformat(),
                    "draw_number": record.draw_number or "",
                    "number": record.number,
                    "source": record.source,
                }
            )


def _record_equals(left: DrawRecord, right: DrawRecord) -> bool:
    return (
        left.draw_type == right.draw_type
        and left.draw_date == right.draw_date
        and left.draw_number == right.draw_number
        and left.digits == right.digits
    )
